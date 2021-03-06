import openpyxl as xl
import json
import random
import enum


def getKanji(level, quantity):
    global wb
    if level == "n5":
        wb = xl.load_workbook('N5_Kanji.xlsx')
    elif level == "n4":
        wb = xl.load_workbook('N4_Kanji_List.xlsx')
    elif level == 'n3':
        wb = xl.load_workbook('N3_Kanji.xlsx')
    sheet = wb['Sheet 1']
    rando = random.sample(range(2, sheet.max_row), int(quantity))
    data = []
    for rand in rando:
        data.append({
            "kanji": sheet.cell(rand, 3).value,
            "hiragana": sheet.cell(rand, 2).value,
            "word": sheet.cell(rand, 1).value
        })

    return data

# def getXKanji(quantity):
#     spl_word = ' '
#     num = int(quantity.partition(spl_word)[2])
#     rando = random.sample(range(2, 332), num)
#     data = []
#     for rand in rando:
#         data.append({
#             "kanji": sheet.cell(rand, 3).value,
#             "hiragana": sheet.cell(rand, 2).value,
#             "word": sheet.cell(rand, 1).value
#         })
#     return data, num
